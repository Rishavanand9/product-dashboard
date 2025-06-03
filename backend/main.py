import pandas as pd
import time
import uuid
import os
import stat
import tempfile
import datetime
import random
import logging
import json
import requests
import platform
from fastapi import FastAPI, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from typing import Dict, List, Optional
from fastapi.middleware.cors import CORSMiddleware
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from PIL import Image as PILImage

# Changes 03062025
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import requests
from io import BytesIO



app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust this to specify allowed origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
log_directory = "logs"
if not os.path.exists(log_directory):
    os.makedirs(log_directory)
log_file = os.path.join(log_directory, f"amazon_scraper_{datetime.datetime.now().strftime('%Y%m%d')}.log")

# Set up logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(lineno)s - %(funcName)20s() - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()  # Keep console output
    ]
)
logger = logging.getLogger(__name__)


def updated_chromedriver(options):

    if platform.system() == "Windows":
        dynamic_link = r"backend\extras\links.json"
    elif platform.system() == "Linux":
        dynamic_link = r"backend/extras/links.json"
    else:
        logger.error("Unsupported OS for chromedriver update")
        raise Exception("Unsupported OS for chromedriver update")
    
    with open(dynamic_link, 'r') as f:
        links = json.load(f)
    driver_link = links.get("chromedriver")
    if not driver_link:
        logger.error("No chromedriver link found in links.json")
        raise Exception("No chromedriver link found in links.json")
    logger.info(f"Updating chromedriver from {driver_link}")
    try:
        response = requests.get(driver_link, timeout=10)
        if response.status_code == 200:
            # Save the driver to a temporary file
            temp_driver_path = os.path.join(tempfile.gettempdir(), "chromedriver")
            with open(temp_driver_path, 'wb') as f:
                f.write(response.content)
            
            # Set executable permissions
            st = os.stat(temp_driver_path)
            os.chmod(temp_driver_path, st.st_mode | stat.S_IEXEC)
            
            # Initialize undetected_chromedriver with the new driver
            driver = uc.Chrome(options=options, browser_executable_path=temp_driver_path)
            logger.info("Successfully updated and initialized undetected_chromedriver")
            return driver
        else:
            logger.error(f"Failed to download chromedriver: Status {response.status_code}")
            raise Exception(f"Failed to download chromedriver: Status {response.status_code}")
    except Exception as e:
        logger.error(f"Error updating chromedriver: {str(e)}")
        raise Exception(f"Error updating chromedriver: {str(e)}")

# Global variables to track job status
active_jobs: Dict[str, Dict] = {}

driver_executable_path = r"D:\FinalProject\product-dashboard\backend\extras\chrome-win64\chrome-win64\chrome.exe"
st = os.stat(driver_executable_path)
os.chmod(driver_executable_path, st.st_mode | stat.S_IEXEC)

def check_file_type(file_path: str):
    """Check the file type of the uploaded file"""
    if file_path.endswith('.xlsx'):
        return "xlsx"
    elif file_path.endswith('.csv'):
        return "csv"
    else:
        return False

def random_sleep(min_seconds=1, max_seconds=3):
    """Random sleep to mimic human behavior and avoid bot detection"""
    sleep_time = random.uniform(min_seconds, max_seconds)
    time.sleep(sleep_time)
    return sleep_time

# Changes 03062025
def download_image(image_url: str, temp_dir: str, index: int, job_id: str) -> Optional[str]:
    """Download an image from a URL and save it temporarily"""
    try:
        response = requests.get(image_url, timeout=10)
        if response.status_code == 200:
            image_path = os.path.join(temp_dir, f"image_{job_id}_{index}.jpg")
            with open(image_path, 'wb') as f:
                f.write(response.content)
            return image_path
        else:
            logger.warning(f"Failed to download image from {image_url}: Status {response.status_code}")
            return None
    except Exception as e:
        logger.error(f"Error downloading image {image_url}: {str(e)}")
        return None

def get_product_info_using_selenium(item_name: str, retry_count: int = 0):
    """Get detailed product information using Selenium with retry mechanism"""
    max_retries = 3
    product_info = {}
    driver = None
    
    try:
        logger.info(f"Searching Amazon for product: {item_name}")
        
        # Configure Chrome options
        options = uc.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        # Initialize undetected chromedriver
        try:
            driver = uc.Chrome(options=options, browser_executable_path=driver_executable_path)
        except Exception as e:
            if str(e).__contains__("This version of ChromeDriver only supports Chrome version"):
                driver = updated_chromedriver(options)
                
            logger.error(f"Failed to initialize undetected_chromedriver: {str(e)}")
            return {"error": f"Failed to initialize browser: {str(e)}"}
        

        # Add random user agent through undetected_chromedriver config
        
        # Go to Amazon.in
        driver.get("https://www.amazon.in")
        
        random_sleep(3, 6)
        
        # Accept cookies if present
        try:
            cookie_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "sp-cc-accept"))
            )
            cookie_button.click()
            random_sleep(1, 3)
        except TimeoutException:
            # Cookie dialog might not appear
            pass
        
        # Search for the product
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "twotabsearchtextbox"))
        )
        search_box.clear()
        # Type like a human - letter by letter with random delays
        for char in item_name:
            search_box.send_keys(char)
            time.sleep(random.uniform(0.05, 0.2))
        
        search_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "nav-search-submit-button"))
        )
        search_button.click()
        
        random_sleep(3, 7)
        
        # Click on the first search result
        try:
            first_result = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div.s-result-item[data-component-type='s-search-result'] img"))
            )
            first_result.click()
            
            # Switch to the new tab if opened
            if len(driver.window_handles) > 1:
                driver.switch_to.window(driver.window_handles[1])
                
            random_sleep(4, 8)
            
            # Extract product information
            product_info = extract_product_details(driver)
            
            logger.info(f"Successfully scraped details for: {item_name}")
            logger.info(f"Product info: {json.dumps(product_info)}")

        except Exception as e:
            logger.error(f"Error finding or clicking first result: {str(e)}")
            product_info["error"] = f"Failed to find product results: {str(e)}"
            
    except Exception as e:
        error_msg = f"Selenium error: {str(e)}"
        logger.error(f"{error_msg} for product: {item_name}")
        product_info["error"] = error_msg
        
        # Retry logic
        if retry_count < max_retries:
            retry_delay = (retry_count + 1) * 5  # Exponential backoff
            logger.info(f"Retrying in {retry_delay} seconds (attempt {retry_count + 1}/{max_retries})...")
            time.sleep(retry_delay)
            if driver:
                driver.quit()
            return get_product_info_using_selenium(item_name, retry_count + 1)
            
    finally:
        # Close the browser
        if driver:
            driver.quit()
            
    return product_info

def extract_product_details(driver):
    """Extract product details from the product page"""
    product_info = {
        "title": "",
        "description": "",
        "image": "",
       "image_urls": [],  # Changed from 'image' to 'image_urls' for multiple images
        "url": driver.current_url,
        "price": "",
        "composition": "",
        "discontinued": "",
        "unspsc_code": "",
        "dimensions": "",
        "weight": "",
        "manufacturer": "",
        "asin": "",
        "model_number": "",  
        "country_of_origin": "",
        "date_first_available": "",
        "included_components": "",
        "generic_name": ""
    }
    
    # Extract title
    try:
        product_info["title"] = driver.find_element(By.ID, "productTitle").text.strip()
        if product_info["title"] == "":
            product_info["title"] = driver.find_element(By.ID, "title").text.strip()
    except NoSuchElementException:
        product_info["title"] = "NA"
    
    # Extract price
    try:
        price_element = driver.find_element(By.CSS_SELECTOR, ".a-price .a-offscreen")
        product_info["price"] = price_element.get_attribute("innerHTML").strip()
    except NoSuchElementException:
        try:
            price_element = driver.find_elements(By.CLASS_NAME, "a-price-whole")[0]
            product_info["price"] = price_element.text.strip().replace("\n.", "")
        except NoSuchElementException:
            product_info["price"] = "NA"
    
    # Extract image URL
    try:
        img_element = driver.find_element(By.ID, "landingImage")
        product_info["image"] = img_element.get_attribute("src")
    except NoSuchElementException:
        product_info["image"] = "NA"


    # Extract multiple image URLs
    try:
        # Target the image thumbnail carousel
        img_elements = driver.find_elements(By.CSS_SELECTOR, ".imageThumbnail img, #altImages img, #main-image-container img")
        image_urls = []
        for img in img_elements[:5]:  # Limit to 5 images to avoid overwhelming
            src = img.get_attribute("src")
            if src and src not in image_urls:
                image_urls.append(src)
        product_info["image_urls"] = image_urls if image_urls else ["NA"]
    except NoSuchElementException:
        product_info["image_urls"] = ["NA"]
    
    # Extract product description
    try:
        description_element = driver.find_element(By.ID, "productDescription")
        product_info["description"] = description_element.text.strip()
    except NoSuchElementException:
        try:
            description_element = driver.find_element(By.CSS_SELECTOR, "#feature-bullets .a-list-item")
            product_info["description"] = "\n".join([item.text for item in driver.find_elements(By.CSS_SELECTOR, "#feature-bullets .a-list-item")])
        except NoSuchElementException:
            product_info["description"] = "NA"
    
    # Extract technical details from product information section
    try:
        # Look for the product details table
        detail_sections = driver.find_elements(By.CSS_SELECTOR, "table.a-keyvalue")
        
        # searching for the product details table in additional information section
        for section in detail_sections:
            rows = section.find_elements(By.CSS_SELECTOR, "tr")
            for row in rows:
                try:
                    key = row.find_element(By.CSS_SELECTOR, "th").text.strip().lower()
                    value = row.find_element(By.CSS_SELECTOR, "td").text.strip()
                    
                    if "asin" in key:
                        product_info["asin"] = value
                    elif "manufacturer" in key:
                        product_info["manufacturer"] = value
                    elif "country of origin" in key:
                        product_info["country_of_origin"] = value
                    elif "date first available" in key:
                        product_info["date_first_available"] = value
                    elif "model" in key and "number" in key:
                        product_info["model_number"] = value
                    elif "item weight" in key or "weight" in key:
                        product_info["weight"] = value
                    elif "dimension" in key:
                        product_info["dimensions"] = value
                    elif "included" in key and "component" in key:
                        product_info["included_components"] = value
                    elif "generic name" in key:
                        product_info["generic_name"] = value
                    elif "composition" in key or "ingredients" in key:
                        product_info["composition"] = value
                except NoSuchElementException:
                    continue
        
        # searching for the product details table in product details section
        detail_list = driver.find_element(By.ID, "detailBullets_feature_div")
        list_items = detail_list.find_elements(By.TAG_NAME, "li")
        detail_sections = "\n".join([item.text.strip() for item in list_items])
        if detail_sections:
            product_info["product_details_as_on_amazon.in"] = detail_sections
        else:
            product_info["product_details_as_on_amazon.in"] = "NA"
                    

    except Exception as e:
        logger.error(f"Error extracting technical details: {str(e)}")
    
    # Check if product is discontinued
    try:
        if "currently unavailable" in driver.page_source.lower() or "we don't know when or if this item will be back in stock" in driver.page_source.lower():
            product_info["discontinued"] = "Yes"
        else:
            product_info["discontinued"] = "No"
    except:
        product_info["discontinued"] = "NA"
        
    return product_info
def download_image_in_memory(image_url: str) -> Optional[BytesIO]:
    """Download an image from a URL and return it as a BytesIO object"""
    try:
        response = requests.get(image_url, timeout=10)
        if response.status_code == 200:
            return BytesIO(response.content)
        else:
            logger.warning(f"Failed to download image from {image_url}: Status {response.status_code}")
            return None
    except Exception as e:
        logger.error(f"Error downloading image {image_url}: {str(e)}")
        return None

def process_file_background(temp_file_path: str, batch_size: int = 5, job_id: str = None):
    try:
        logger.info(f"Starting job {job_id}: Reading file {temp_file_path}")
        if "xlsx" in temp_file_path:
            df = pd.read_excel(temp_file_path)
        elif "csv" in temp_file_path:
            df = pd.read_csv(temp_file_path)
        
        total_products = len(df)
        results = []
        processed = 0
        
        logger.info(f"Job {job_id}: Found {total_products} products to process")
        active_jobs[job_id]["total"] = total_products

        # Process in batches
        for i in range(0, total_products, batch_size):
            batch = df.iloc[i:i+batch_size]
            batch_num = i//batch_size + 1
            total_batches = (total_products-1)//batch_size + 1
            logger.info(f"Job {job_id}: Processing batch {batch_num}/{total_batches}")
            
            for _, row in batch.iterrows():
                srno = row.get("SrNo", "NA")
                item_code = row.get("Item Code", "NA")
                name = row.get("Item Name")
                
                if not name:
                    logger.info(f"Job {job_id}: Skipping empty product name")
                    processed += 1
                    continue

                # Get product info using Selenium
                product_info = get_product_info_using_selenium(name)

                # Create a flattened result dictionary with all the information
                result = {
                    "SrNo": srno,
                    "Item Code": item_code,
                    "Item Name": name,
                    "Title": product_info.get("title", ""),
                    "Composition_on_amazon.in": product_info.get("description", ""),
                    "Price": product_info.get("price", ""),
                    "Product Details as on amazon.in": product_info.get("product_details_as_on_amazon.in", ""),
                    "Image URL": product_info.get("image", ""),
                    "Image URLs": product_info.get("image_urls", []),  # List for multiple images
                    "Amazon URL": product_info.get("url", ""),
                    "Is Discontinued": product_info.get("discontinued", ""),
                    "UNSPSC Code": product_info.get("unspsc_code", ""),
                    "Product Dimensions": product_info.get("dimensions", ""),
                    "Item Weight": product_info.get("weight", ""),
                    "Manufacturer": product_info.get("manufacturer", ""),
                    "ASIN": product_info.get("asin", ""),
                    "Model Number": product_info.get("model_number", ""),
                    "Country of Origin": product_info.get("country_of_origin", ""),
                    "Date First Available": product_info.get("date_first_available", ""),
                    "Included Components": product_info.get("included_components", ""),
                    "Generic Name": product_info.get("generic_name", ""),
                    "Error": product_info.get("error", "")
                }
                
                results.append(result)
                
                processed += 1
                active_jobs[job_id]["processed"] = processed
                progress_pct = (processed/total_products*100)
                logger.info(f"Job {job_id}: Progress {processed}/{total_products} ({progress_pct:.1f}%)")
                
                wait_time = random.uniform(10, 20)
                logger.info(f"Job {job_id}: Waiting {wait_time:.1f} seconds before next item")
                time.sleep(wait_time)
            
            if i + batch_size < total_products:
                pause_time = random.uniform(30, 60)
                logger.info(f"Job {job_id}: Batch {batch_num} complete. Pausing for {pause_time:.1f} seconds before next batch")
                time.sleep(pause_time)
        
        logger.info(f"Job {job_id}: All products processed. Saving results to output")
        output_dir = os.path.join(os.path.dirname(__file__), "output_files")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_path = os.path.join(output_dir, f"output_{job_id}.xlsx")
        
        if temp_file_path.endswith('.xlsx'):
            # Save to Excel
            df_results = pd.DataFrame(results)
            df_results.to_excel(output_path, index=False)
            
            # Load workbook to add images
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Adjust column width and row height
            ws.column_dimensions['I'].width = 50  # Image URLs column
            ws.column_dimensions['J'].width = 20  # Embedded Images column
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 100  # Adjust row height for images
            
            # Add header for image column
            ws['J1'] = "Embedded Images"
            
            # Embed images
            for row_idx, result in enumerate(results, start=2):
                if len(result["Image URLs"]) == 1 and result["Image URLs"][0] == "NA":
                    logger.info(f"Job {job_id}: No images to embed for row {row_idx}")
                    continue
                image_urls = result["Image URLs"]
                for idx, url in enumerate(image_urls[:5]):  # Limit to 5 images
                    if url != "NA":
                        image_data = download_image_in_memory(url)
                        if image_data:
                            try:
                                # Open image with PIL and convert to a format openpyxl can use
                                pil_img = PILImage.open(image_data)
                                # Resize image (optional, adjust as needed)
                                # pil_img = pil_img.resize((100, 100), PILImage.Resampling.LANCZOS)
                                # Save to BytesIO in a compatible format (PNG)
                                img_buffer = BytesIO()
                                pil_img.save(img_buffer, format="PNG")
                                img_buffer.seek(0)
                                # Embed in Excel
                                img = OpenpyxlImage(img_buffer)
                                ws.add_image(img, f"J{row_idx}")
                            except Exception as e:
                                logger.error(f"Failed to embed image {url} for row {row_idx}: {str(e)}")
            
            wb.save(output_path)
        elif temp_file_path.endswith('.csv'):
            pd.DataFrame(results).to_csv(output_path.replace('.xlsx', '.csv'), index=False)
            logger.warning(f"Job {job_id}: CSV output does not support image embedding")

        active_jobs[job_id]["status"] = "completed"
        active_jobs[job_id]["output_file"] = output_path
        logger.info(f"Job {job_id}: Job completed successfully")
        
    except Exception as e:
        error_msg = f"Error processing Excel: {str(e)}"
        logger.error(f"{error_msg}")
        active_jobs[job_id]["status"] = "failed"
        active_jobs[job_id]["error"] = error_msg
    finally:
        # Clean up the temporary file
        try:
            os.unlink(temp_file_path)
            logger.info(f"Job {job_id}: Temporary file cleaned up")
        except Exception as e:
            logger.error(f"Job {job_id}: Failed to cleanup temporary file: {str(e)}")

@app.post("/upload/")
async def upload_file(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    """Upload Excel/CSV file with product names for processing"""
    filetype = ""
    try:
        logger.info(f"Received file upload: {file.filename}")
        filetype = check_file_type(file.filename)
        if not filetype:
            return JSONResponse(status_code=400, content={"error": "Unsupported file type"})

        # Create a temporary file to store the uploaded content
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{filetype}') as temp_file:
            # Write the file content to the temporary file
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        # Generate a job ID
        job_id = uuid.uuid4().hex[:8]
        
        # Initialize job status
        active_jobs[job_id] = {
            "status": "processing",
            "processed": 0,
            "total": 0,
            "start_time": datetime.datetime.now().isoformat(),
            "file_name": file.filename
        }

        logger.info(f"Active jobs: {active_jobs}")
        logger.info(f"New job created: {job_id} for file {file.filename}")
        
        # Start processing in background
        background_tasks.add_task(process_file_background, temp_file_path, batch_size=5, job_id=job_id)
        
        return {
            "job_id": job_id, 
            "message": "Processing started. Use /status/{job_id} to check progress and /download/{job_id} to get results when complete"
        }
    except Exception as e:
        error_msg = f"Error handling upload: {str(e)}"
        logger.error(f"{error_msg}")
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.get("/status/{job_id}")
async def get_job_status(job_id: str):
    """Check the status of a processing job"""
    if job_id not in active_jobs:
        return JSONResponse(status_code=404, content={"error": "Job not found"})
    
    job_info = active_jobs[job_id].copy()
    
    # Calculate and add progress percentage
    if job_info["total"] > 0:
        job_info["progress_percentage"] = round((job_info["processed"] / job_info["total"]) * 100, 1)
    else:
        job_info["progress_percentage"] = 0
        
    # Calculate elapsed time
    start_time = datetime.datetime.fromisoformat(job_info["start_time"])
    elapsed = (datetime.datetime.now() - start_time).total_seconds()
    job_info["elapsed_seconds"] = elapsed
    job_info["elapsed_formatted"] = str(datetime.timedelta(seconds=int(elapsed)))
    
    return job_info

@app.get("/download/{job_id}")
async def download_results(job_id: str):
    """Download the results of a completed job"""
    if job_id not in active_jobs:
        return JSONResponse(status_code=404, content={"error": "Job not found"})
    
    job = active_jobs[job_id]
    if job["status"] != "completed":
        return JSONResponse(
            status_code=400, 
            content={
                "error": f"Job is not completed. Current status: {job['status']}",
                "processed": job["processed"],
                "total": job["total"],
                "progress_percentage": round((job["processed"] / job["total"]) * 100, 1) if job["total"] > 0 else 0
            }
        )
    
    output_file = job.get("output_file")
    if not output_file or not os.path.exists(output_file):
        return JSONResponse(status_code=404, content={"error": "Output file not found"})
    
    logger.info(f"Sending results file for job {job_id}: {output_file}")
    return FileResponse(output_file, filename=f"amazon_results_{job_id}.xlsx")

@app.get("/jobs")
async def list_jobs():
    """List all active and completed jobs"""
    job_summaries = {}
    for job_id, job_info in active_jobs.items():
        job_summaries[job_id] = {
            "status": job_info["status"],
            "file_name": job_info.get("file_name", "Unknown"),
            "processed": job_info["processed"],
            "total": job_info["total"],
            "progress_percentage": round((job_info["processed"] / job_info["total"]) * 100, 1) if job_info["total"] > 0 else 0,
            "start_time": job_info["start_time"]
        }
    return job_summaries